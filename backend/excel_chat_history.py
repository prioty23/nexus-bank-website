from openpyxl import Workbook, load_workbook #W-create new excel & L- Opens exsisting excel
from datetime import datetime #chat can save each timestamp
import os #check excel file exists


EXCEL_FILE = "chat_logs.xlsx" #stores excel file with name


def create_excel_file_if_not_exists(): #excel file already exists or if not then creates 
    if not os.path.exists(EXCEL_FILE):  #not found create new one
        workbook = Workbook() #creates a brand new excel file
        sheet = workbook.active #selects the first page inside that Excel file so the code can save & read data there.
        sheet.title = "Chat Logs"

        sheet.append([ #add 1st row and columns names
            "session_id",
            "complaint_id",
            "user_message",
            "bot_reply",
            "source",
            "blocked",
            "status",
            "created_at"
        ])

        workbook.save(EXCEL_FILE)  #saves the new excel file


def save_chat_to_excel(  #saves each chat & reply into excel
    session_id,
    user_message,
    bot_reply,
    source,
    blocked,
    status,
    complaint_id=""
):
    create_excel_file_if_not_exists() #makes sure the Excel file exists before writing into it.

    workbook = load_workbook(EXCEL_FILE) #opens existing file
    sheet = workbook.active #

    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S") #gets current date time

    sheet.append([ #adds new row with current chat data
        session_id,
        complaint_id,
        user_message,
        bot_reply,
        source,
        blocked,
        status,
        created_at
    ])

    workbook.save(EXCEL_FILE) #saves the file after adding new row


def get_chat_history_from_excel(session_id, limit=6): #reads old tetx from excel
    create_excel_file_if_not_exists() #makes sure the Excel file exists before reading it.

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    history = [] #creates an empty list to store previous text

    for row in sheet.iter_rows(min_row=2, values_only=True): #reads from row2 row1 is header
        row_session_id = row[0]
        user_message = row[2]
        bot_reply = row[3]
        blocked = row[5]

        if row_session_id == session_id and blocked == False:  #only reads history from the same user/session, were not blocked
            if user_message:
                history.append({
                    "role": "user",
                    "content": user_message #text from user
                })

            if bot_reply:
                history.append({
                    "role": "assistant",
                    "content": bot_reply #reply from bot
                })

    return history[-limit:]  #returns only latest 6 texts